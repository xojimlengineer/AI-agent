SYSTEM_PROMPT = """
# Siz — bank ma'lumotlar bazasiga SQL yozadigan yordamchisiz.
#
# Baza sxemasi:
# clients (id, name, birth_date, region)
# accounts (id, client_id, balance, open_date)
# transactions (id, account_id, amount, date, type)
#
# QOIDALAR:
# - Faqat SELECT yozasiz. DDL/DML yo‘q.
# - Region filtri clients.region ustunida.
# - Region mapping (UZ → DB):
#   'Toshkent' → 'Tashkent'
#   'Jizzax' → 'Jizzakh'
#   'Samarqand' → 'Samarkand'
#   'Andijon' → 'Andijan'
#   'Farg‘ona' → 'Fergana'
#   'Namangan' → 'Namangan'
#   'Buxoro' → 'Bukhara'
#   'Xorazm' → 'Khorezm'
#   'Qashqadaryo' → 'Qashqadarya'
#   'Surxondaryo' → 'Surxondaryo'
#   'Navoiy' → 'Navoi'
#   'Sirdaryo' → 'Sirdarya'
#   'Qoraqalpog‘iston' → 'Karakalpakstan'
#
# - O‘zbek oy nomlari → raqam:
#   yanvar=1, fevral=2, mart=3, aprel=4, may=5, iyun=6, iyul=7, avgust=8, sentabr=9, oktabr=10, noyabr=11, dekabr=12.
#
# - Bir oy bir nechta yilda so‘ralsa (mas: "2023 va 2024 yil iyun"):
#   Agar umumiy yig‘indi bo‘lsa:
#     SELECT SUM(t.amount) AS total_amount
#     FROM transactions t
#     JOIN accounts a ON t.account_id = a.id
#     JOIN clients  c ON a.client_id = c.id
#     WHERE EXTRACT(MONTH FROM t.date) = <oy>
#       AND EXTRACT(YEAR FROM t.date) IN (<yillar>)
#       AND c.region = '<DB_region>';
#   Agar "har yil bo‘yicha" deyilgan bo‘lsa:
#     SELECT EXTRACT(YEAR FROM t.date) AS year, SUM(t.amount) AS total_amount
#     FROM ...
#     WHERE EXTRACT(MONTH FROM t.date) = <oy>
#       AND EXTRACT(YEAR FROM t.date) IN (<yillar>)
#       AND c.region = '<DB_region>'
#     GROUP BY 1 ORDER BY 1;
#
# - Bir nechta oylar so‘ralsa (mas: "2023 yil may, iyun va iyul" yoki "maydan avgustgacha"):
#   Oylar ro‘yxatini raqam qilib oling va IN (...) bilan filtrlang.
#   * Agar "har oy bo‘yicha / oylar kesimida / month-wise" bo‘lsa:
#       SELECT EXTRACT(YEAR FROM t.date) AS year,
#              EXTRACT(MONTH FROM t.date) AS month,
#              SUM(t.amount) AS total_amount
#       FROM ...
#       WHERE EXTRACT(YEAR FROM t.date) IN (<yillar>)
#         AND EXTRACT(MONTH FROM t.date) IN (<oylar>)
#         AND c.region = '<DB_region>'
#       GROUP BY 1,2 ORDER BY 1,2;
#   * Aks holda umumiy SUM qaytar:
#       SELECT SUM(t.amount) AS total_amount
#       FROM ...
#       WHERE EXTRACT(YEAR FROM t.date) IN (<yillar>)
#         AND EXTRACT(MONTH FROM t.date) IN (<oylar>)
#         AND c.region = '<DB_region>';
#
# MUHIM:
# - DBga doir savollarda har doim YAGONA SQL yozing va DARHOL `assistant_excel` tool’ini chaqiring.
# - DB bilan aloqasi bo‘lmagan savollarda oddiy matn qaytaring.
# """

import os
from typing import TypedDict, Annotated, Sequence, List, Dict
from operator import add as add_messages
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv

from langgraph.graph import StateGraph, END
from langchain_core.messages import BaseMessage, SystemMessage, ToolMessage
from langchain_openai import ChatOpenAI
from langchain_core.tools import tool
from langchain_core.runnables import RunnableConfig

from sqlalchemy import create_engine, text

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

load_dotenv()

DB_NAME = os.getenv("DB_NAME", "")
DB_USER = os.getenv("DB_USER", "")
DB_PASS = os.getenv("DB_PASS", "")
DB_HOST = os.getenv("DB_HOST", "127.0.0.1")
DB_PORT = os.getenv("DB_PORT", "5432")

MODEL_OPENAI = os.getenv("MODEL_OPENAI")
API_KEY_OPENAI = os.getenv("API_KEY_OPENAI")
# SYSTEM_PROMPT = os.getenv("SYSTEM_PROMPT", "Siz SQL yozib, tool orqali bajaradigan agentsiz.")

# ---- PostgreSQL engine
ENGINE = create_engine(
    f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}",
    pool_pre_ping=True,
)

def save_with_chart(df: pd.DataFrame) -> str:
    fp = (Path(__file__).parent / "result.xlsx").as_posix()
    wb, ws = Workbook(), Workbook().active  # placeholder to satisfy type checkers
    wb = Workbook(); ws = wb.active

    # 1x1 bo‘lsa grafik uchun label/valuega aylantiramiz
    if df.shape == (1, 1):
        df = pd.DataFrame({"label": [df.columns[0]], "value": [df.iloc[0, 0]]})

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    num_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    if num_cols:
        first_c = df.columns.get_loc(num_cols[0]) + 1
        last_c = df.columns.get_loc(num_cols[-1]) + 1

        data = Reference(ws, min_row=1, max_row=ws.max_row, min_col=first_c, max_col=last_c)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

        bar = BarChart(); bar.varyColors = True
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True

        if ws.max_row > 3:
            line = LineChart()
            line.add_data(data, titles_from_data=True)
            line.set_categories(cats)
            bar += line

        ws.add_chart(bar, f"{get_column_letter(ws.max_column + 3)}2")

    wb.save(fp)
    return fp

@tool
def assistant_excel(prompt: str):
    """SELECT SQL ni Postgres’da bajarib, natijani Excelga saqlaydi."""
    with ENGINE.begin() as conn:
        df = pd.read_sql(text(prompt), conn)
    filename = save_with_chart(df)
    return {
        "reply": "Excel tayyor",
        "_file": {
            "path": filename,
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "caption": "file"
        }
    }

tools = [assistant_excel]
llm = ChatOpenAI(model=MODEL_OPENAI, temperature=0, api_key=API_KEY_OPENAI).bind_tools(tools)

def concat_list(prev, new):
    return (prev or []) + (new or [])

class AgentState(TypedDict):
    messages: Annotated[Sequence[BaseMessage], add_messages]
    outbox: Annotated[List[Dict], concat_list]

def should_continue(state: AgentState):
    last = state["messages"][-1]
    return hasattr(last, "tool_calls") and len(last.tool_calls) > 0

tools_dict = {t.name: t for t in tools}

def call_llm(state: AgentState, config: RunnableConfig):
    msgs = [SystemMessage(content=SYSTEM_PROMPT)] + state["messages"]
    resp = llm.invoke(msgs)
    return {"messages": [resp]}

def take_action(state: AgentState, config: RunnableConfig) -> AgentState:
    tool_calls = state["messages"][-1].tool_calls
    out_files: List[Dict] = []

    for t in tool_calls:
        name, args = t["name"], t["args"]
        if name not in tools_dict:
            state["messages"].append(ToolMessage(tool_call_id=t["id"], name=name, content="Unknown tool"))
            continue
        res = tools_dict[name].invoke({**args})
        reply_text = res.get("reply") if isinstance(res, dict) else str(res)
        if isinstance(res, dict) and "_file" in res:
            out_files.append(res["_file"])
        state["messages"].append(ToolMessage(tool_call_id=t["id"], name=name, content=str(reply_text)))

    if out_files:
        state["outbox"] = out_files
    return state

graph = StateGraph(AgentState)
graph.add_node("llm", call_llm)
graph.add_node("retriever_agent", take_action)
graph.add_conditional_edges("llm", should_continue, {True: "retriever_agent", False: END})
graph.add_edge("retriever_agent", "llm")
graph.set_entry_point("llm")

rag_agent = graph.compile()
